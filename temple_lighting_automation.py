#!/usr/bin/env python3
"""
Temple Lighting Donation Automation Tool (点灯自动化工具)
=====================================================

This script automates the monthly lighting donation spreadsheet processing.
Run this twice per month for initial day (初一) and 15th day (十五日).

Usage:
    python temple_lighting_automation.py

Requirements:
    - openpyxl
    - pillow (for image operations if needed)

Install dependencies:
    pip install openpyxl pillow reportlab
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import os
from pathlib import Path
from datetime import datetime
import json
import sys


class TempleWorkflow:
    def __init__(self, excel_path, year_stem="丙午"):
        """Initialize the workflow with Excel file path"""
        self.excel_path = Path(excel_path)
        self.year_stem = year_stem
        if not self.excel_path.exists():
            print(f"❌ Error: File not found at {excel_path}")
            sys.exit(1)
        self.wb = openpyxl.load_workbook(excel_path)
        print(f"✅ Loaded workbook: {self.excel_path.name}")

    def clean_names(self, text):
        """Extract clean names from pasted text with numbers and checkmarks"""
        lines = text.strip().split('\n')
        names = []
        for line in lines:
            # Remove number prefix (1., 2., etc.)
            cleaned = re.sub(r'^[\d\.]+\s*⁠?\s*', '', line)
            # Remove invisible characters (\u2060), checkmarks, extra spaces
            cleaned = re.sub(r'[\u2060✔️✓\s]+', ' ', cleaned).strip()
            if cleaned and len(cleaned) > 0:
                names.append(cleaned)
        return names

    def load_names_from_file(self, file_path):
        """Load names from a text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return self.clean_names(f.read())
        except FileNotFoundError:
            print(f"❌ Names file not found: {file_path}")
            return []

    def load_names_from_input(self):
        """Interactive input for names"""
        print("\n" + "="*60)
        print("PASTE NAMES LIST")
        print("="*60)
        print("Paste your name list (with numbering and checkmarks).")
        print("Press Enter twice when done:")
        lines = []
        empty_count = 0
        while True:
            try:
                line = input()
                if line == "":
                    empty_count += 1
                    if empty_count >= 2:
                        break
                else:
                    empty_count = 0
                    lines.append(line)
            except EOFError:
                break
        return self.clean_names('\n'.join(lines))

    def clear_sheet_data(self, sheet_name):
        """Clear old data from the sheet"""
        ws = self.wb[sheet_name]
        # Clear names columns (B, E, H, K)
        name_cols = [2, 5, 8, 11]
        for col in name_cols:
            for row in range(5, 31):
                ws.cell(row=row, column=col).value = None
        # Clear numbering (column A)
        for row in range(5, 31):
            ws.cell(row=row, column=1).value = None
        print(f"✅ Cleared old data from {sheet_name}")

    def insert_names(self, sheet_name, names):
        """Insert names into sheet columns B, E, H, K starting from row 5"""
        ws = self.wb[sheet_name]
        name_cols = [2, 5, 8, 11]  # B, E, H, K
        
        name_counter = 0
        for col in name_cols:
            row = 5
            row_num = 1
            while row <= 30 and name_counter < len(names):
                ws.cell(row=row, column=1).value = row_num
                ws.cell(row=row, column=col).value = names[name_counter]
                row_num += 1
                row += 1
                name_counter += 1
        
        print(f"✅ Inserted {len(names)} names into {sheet_name}")
        return name_counter

    def format_names_font(self, sheet_name, font_name='KaiTi TC', font_size=15):
        """Apply font formatting to all names"""
        ws = self.wb[sheet_name]
        font = Font(name=font_name, size=font_size)
        alignment = Alignment(horizontal='center', vertical='center')
        
        name_cols = [2, 5, 8, 11]
        for col in name_cols:
            for row in range(5, 31):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.font = font
                    cell.alignment = alignment
        
        # Format row numbers
        number_font = Font(name=font_name, size=15)
        for row in range(5, 31):
            cell = ws.cell(row=row, column=1)
            if cell.value:
                cell.font = number_font
                cell.alignment = alignment
        
        print(f"✅ Applied KaiTi TC font (size 15) to all names")

    def add_borders(self, sheet_name):
        """Add borders to match template style"""
        ws = self.wb[sheet_name]
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to name cells
        name_cols = [2, 5, 8, 11]
        for col in name_cols:
            for row in range(5, 31):
                ws.cell(row=row, column=col).border = thin_border
        
        print(f"✅ Applied borders to {sheet_name}")

    def update_date(self, sheet_name, lunar_date):
        """Update date in row 2"""
        ws = self.wb[sheet_name]
        date_text = f"{self.year_stem}年{lunar_date}日"
        ws['A2'] = date_text
        
        # Format date cell
        date_font = Font(name='KaiTi TC', size=16, bold=False)
        ws['A2'].font = date_font
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"✅ Updated date to: {date_text}")

    def add_footer(self, sheet_name, footer_text="出入平安   生意興隆   身體健康"):
        """Add footer text at bottom"""
        ws = self.wb[sheet_name]
        
        # Find the last row with data
        max_row = 35
        
        ws[f'A{max_row}'] = footer_text
        footer_font = Font(name='KaiTi TC', size=38, bold=True)
        ws[f'A{max_row}'].font = footer_font
        ws[f'A{max_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"✅ Added footer text")

    def save_workbook(self):
        """Save the modified workbook"""
        self.wb.save(self.excel_path)
        print(f"✅ Saved workbook: {self.excel_path}")

    def export_to_pdf(self, output_dir, month_date):
        """Export the sheet to PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            print(f"⚠️  PDF export requires LibreOffice. Please export manually from Excel.")
            print(f"   File saved at: {self.excel_path}")
        except ImportError:
            print(f"⚠️  PDF export skipped. reportlab not installed.")
            print(f"   To enable: pip install reportlab")

    def list_available_sheets(self):
        """List all available month sheets"""
        month_sheets = [name for name in self.wb.sheetnames 
                       if name not in ['Sheet1', '甲辰年乐捐-八爷签收', '頭牙', '沈府尾牙']]
        return month_sheets

    def run_workflow(self, sheet_name, names, lunar_date):
        """Execute the complete workflow"""
        print("\n" + "="*60)
        print(f"PROCESSING: {sheet_name}")
        print("="*60)
        
        # Step 1: Clear old data
        self.clear_sheet_data(sheet_name)
        
        # Step 2: Insert names
        self.insert_names(sheet_name, names)
        
        # Step 3: Format fonts
        self.format_names_font(sheet_name)
        
        # Step 4: Add borders
        self.add_borders(sheet_name)
        
        # Step 5: Update date
        self.update_date(sheet_name, lunar_date)
        
        # Step 6: Add footer
        self.add_footer(sheet_name)
        
        # Step 7: Save
        self.save_workbook()
        
        print("\n" + "="*60)
        print("✅ WORKFLOW COMPLETE!")
        print("="*60)
        print(f"Updated sheet: {sheet_name}")
        print(f"Total names processed: {len(names)}")
        print(f"File saved: {self.excel_path}")


def main():
    """Main entry point"""
    print("\n" + "="*60)
    print("🏯 TEMPLE LIGHTING DONATION AUTOMATION TOOL")
    print("="*60)
    
    # Get Excel file path
    print("\nEnter the path to your Excel file:")
    print("  Example: /Users/username/Desktop/temple/点灯/丙午年初一十五点灯.xlsx")
    excel_path = input("File path: ").strip()
    
    if not excel_path:
        print("❌ No file path provided")
        return
    
    # Initialize workflow
    workflow = TempleWorkflow(excel_path)
    
    # List available sheets
    print("\n📅 Available month sheets:")
    sheets = workflow.list_available_sheets()
    for i, sheet in enumerate(sheets, 1):
        print(f"  {i}. {sheet}")
    
    # Select sheet
    print("\nWhich sheet do you want to process? (e.g., '三月十五日' or just '3')")
    sheet_choice = input("Enter sheet name or number: ").strip()
    
    # Parse sheet choice
    if sheet_choice.isdigit():
        sheet_idx = int(sheet_choice) - 1
        if 0 <= sheet_idx < len(sheets):
            sheet_name = sheets[sheet_idx]
        else:
            print("❌ Invalid selection")
            return
    else:
        sheet_name = sheet_choice
        if sheet_name not in sheets:
            print(f"❌ Sheet '{sheet_name}' not found")
            return
    
    print(f"\nSelected: {sheet_name}")
    
    # Input method for names
    print("\nHow do you want to input names?")
    print("  1. Paste from clipboard")
    print("  2. Load from file")
    choice = input("Choice (1 or 2): ").strip()
    
    if choice == "2":
        file_path = input("Enter names file path: ").strip()
        names = workflow.load_names_from_file(file_path)
    else:
        names = workflow.load_names_from_input()
    
    if not names:
        print("❌ No names provided")
        return
    
    print(f"\n✅ Loaded {len(names)} names")
    
    # Extract lunar date from sheet name
    # Examples: "三月十五日" -> "三月十五"
    lunar_date = sheet_name.replace('日', '')
    
    # Run the workflow
    workflow.run_workflow(sheet_name, names, lunar_date)


if __name__ == "__main__":
    main()
