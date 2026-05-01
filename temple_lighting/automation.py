#!/usr/bin/env python3
"""
Temple Lighting Donation Automation Tool (点灯自动化工具)
=====================================================

Automates the monthly lighting donation spreadsheet processing.
Run twice per month for initial day (初一) and 15th day (十五日).

Usage:
    python -m temple_lighting.automation

Requirements:
    openpyxl
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import os
from pathlib import Path
from datetime import datetime
import json
import sys


class TempleWorkflow:
    def __init__(self, excel_path, year_stem="丙午"):
        """Initialize the workflow with Excel file path"""
        self.excel_path = Path(excel_path)
        self.year_stem = year_stem
        self._local_config = _load_local_config()
        if not self.excel_path.exists():
            print(f"❌ Error: File not found at {excel_path}")
            sys.exit(1)
        self.wb = openpyxl.load_workbook(excel_path)
        print(f"✅ Loaded workbook: {self.excel_path.name}")

    def is_english(self, text):
        """Return True if text contains ASCII letters (i.e. an English name)."""
        return bool(re.search(r'[A-Za-z]', text))

    def clean_names(self, text):
        """Extract clean names from pasted text with numbers and checkmarks"""
        lines = text.strip().split('\n')
        names = []
        for line in lines:
            # Remove optional leading whitespace + number prefix (1., 2., etc.)
            # and any trailing invisible chars / spaces after the number
            cleaned = re.sub(r'^\s*\d+[.。]?\s*[\u2060\s]*', '', line)
            # Remove any remaining checkmarks and invisible characters
            cleaned = re.sub(r'[\u2060✔️✓]+', '', cleaned).strip()
            if cleaned:
                names.append(cleaned)
        return names

    def load_names_from_file(self, file_path):
        """Load names from a text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return self.clean_names(f.read())
        except FileNotFoundError:
            print(f"❌ Names file not found: {file_path}")
            return []

    def load_names_from_input(self, first_line=None):
        """Interactive input for names. first_line rescues a line already consumed by a prior prompt."""
        print("\n" + "="*60)
        print("PASTE NAMES LIST")
        print("="*60)
        print("Paste your name list (with numbering and checkmarks).")
        print("Press Enter twice when done:")
        lines = []
        if first_line:
            lines.append(first_line)
        empty_count = 0
        while True:
            try:
                line = input()
                if line == "":
                    empty_count += 1
                    if empty_count >= 2:
                        break
                else:
                    empty_count = 0
                    lines.append(line)
            except EOFError:
                break
        return self.clean_names('\n'.join(lines))

    def clear_sheet_data(self, sheet_name):
        """Clear old data and borders from the sheet (also removes template border typos)."""
        ws = self.wb[sheet_name]
        all_cols = [1, 2, 4, 5, 7, 8, 10, 11]
        for col in all_cols:
            for row in range(5, 31):
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell.border = Border()  # wipe any stray template borders
        print(f"✅ Cleared old data from {sheet_name}")

    def insert_names(self, sheet_name, names):
        """Insert names into columns B, E, H, K with sequential numbering in A, D, G, J"""
        ws = self.wb[sheet_name]
        name_cols = [2, 5, 8, 11]   # B, E, H, K
        num_cols  = [1, 4, 7, 10]   # A, D, G, J  (one number column per name column)

        name_counter = 0
        for name_col, num_col in zip(name_cols, num_cols):
            row = 5
            while row <= 30 and name_counter < len(names):
                ws.cell(row=row, column=num_col).value = name_counter + 1  # sequential
                ws.cell(row=row, column=name_col).value = names[name_counter]
                row += 1
                name_counter += 1

        print(f"✅ Inserted {len(names)} names into {sheet_name}")
        return name_counter

    def format_names_font(self, sheet_name, font_name='KaiTi TC', font_size=15):
        """Apply font formatting to all names"""
        ws = self.wb[sheet_name]

        chinese_font  = Font(name=font_name, size=font_size)
        english_font  = Font(name='Aptos', size=13)
        number_font   = Font(name=font_name, size=font_size)
        left_align    = Alignment(horizontal='left',   vertical='center')
        shrink_align  = Alignment(horizontal='left',   vertical='center', shrink_to_fit=True)
        center_align  = Alignment(horizontal='center', vertical='center')

        name_cols = [2, 5, 8, 11]   # B, E, H, K
        num_cols  = [1, 4, 7, 10]   # A, D, G, J

        for col in name_cols:
            for row in range(5, 31):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    if self.is_english(str(cell.value)):
                        cell.font      = english_font
                        cell.alignment = shrink_align
                    else:
                        cell.font      = chinese_font
                        cell.alignment = left_align

        for col in num_cols:
            for row in range(5, 31):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.font      = number_font
                    cell.alignment = center_align

        print(f"✅ Applied font formatting to all names")

    def add_borders(self, sheet_name):
        """Outer edges of each section: solid thin. Inner cell dividers: dashed.
        Only borders rows that contain data — empty trailing rows get no border."""
        ws = self.wb[sheet_name]

        solid  = Side(style='thin')
        dashed = Side(style='dashed')
        sections = [(1, 2), (4, 5), (7, 8), (10, 11)]

        for num_col, name_col in sections:
            # Find the last row that has a name in this section
            last_row = None
            for row in range(30, 4, -1):
                if ws.cell(row=row, column=name_col).value:
                    last_row = row
                    break

            if last_row is None:
                continue  # no data in this section at all

            for row in range(5, last_row + 1):
                top    = solid if row == 5        else dashed
                bottom = solid if row == last_row else dashed
                ws.cell(row=row, column=num_col).border = Border(
                    left=solid, right=dashed, top=top, bottom=bottom
                )
                ws.cell(row=row, column=name_col).border = Border(
                    left=dashed, right=solid, top=top, bottom=bottom
                )

        print(f"✅ Applied borders to {sheet_name}")

    def update_date(self, sheet_name, lunar_date):
        """Update date in row 2"""
        ws = self.wb[sheet_name]
        date_text = f"{self.year_stem}年{lunar_date}日"
        ws['A2'] = date_text

        # Format date cell
        date_font = Font(name='KaiTi TC', size=24, bold=True)
        ws['A2'].font = date_font
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        print(f"✅ Updated date to: {date_text}")

    def add_footer(self, sheet_name, footer_text="出入平安   生意興隆   身體健康"):
        """Update existing footer cell font. Writes text only if no footer found in template."""
        ws = self.wb[sheet_name]

        # Scan rows 30-45 for existing footer text so we don't create a duplicate
        footer_cell = None
        keywords = ['出入平安', '生意興隆', '身體健康']
        for row in range(30, 46):
            cell = ws.cell(row=row, column=1)
            val = str(cell.value) if cell.value else ''
            if any(kw in val for kw in keywords):
                footer_cell = cell
                break

        if footer_cell is None:
            footer_cell = ws['A35']
            footer_cell.value = footer_text

        footer_cell.font      = Font(name='KaiTi TC', size=32, bold=True)
        footer_cell.alignment = Alignment(horizontal='center', vertical='center')
        print(f"✅ Applied footer formatting at row {footer_cell.row}")

    def save_workbook(self):
        """Save the modified workbook"""
        self.wb.save(self.excel_path)
        print(f"✅ Saved workbook: {self.excel_path}")

    def export_to_pdf(self, output_dir, month_date):
        """Export the sheet to PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            print(f"⚠️  PDF export requires LibreOffice. Please export manually from Excel.")
            print(f"   File saved at: {self.excel_path}")
        except ImportError:
            print(f"⚠️  PDF export skipped. reportlab not installed.")
            print(f"   To enable: pip install reportlab")

    def list_available_sheets(self):
        """List all available month sheets"""
        month_sheets = [name for name in self.wb.sheetnames
                       if name not in ['Sheet1', '甲辰年乐捐-八爷签收', '頭牙', '沈府尾牙']]
        return month_sheets

    def run_workflow(self, sheet_name, names, lunar_date):
        """Execute the complete workflow"""
        print("\n" + "="*60)
        print(f"PROCESSING: {sheet_name}")
        print("="*60)

        # Step 1: Clear old data
        self.clear_sheet_data(sheet_name)

        # Step 2: Insert names
        self.insert_names(sheet_name, names)

        # Step 3: Format fonts
        self.format_names_font(sheet_name)

        # Step 4: Add borders
        self.add_borders(sheet_name)

        # Step 5: Update date
        self.update_date(sheet_name, lunar_date)

        # Step 6: Save (footer is part of the template's print footer — not written to cells)
        self.save_workbook()

        print("\n" + "="*60)
        print("✅ WORKFLOW COMPLETE!")
        print("="*60)
        print(f"Updated sheet: {sheet_name}")
        print(f"Total names processed: {len(names)}")
        print(f"File saved: {self.excel_path}")

        # Open the file automatically
        import subprocess, sys
        if sys.platform == "darwin":
            open_with = self._local_config.get("open_with")
            cmd = ["open", "-a", open_with, str(self.excel_path)] if open_with else ["open", str(self.excel_path)]
            subprocess.run(cmd)
        elif sys.platform == "win32":
            os.startfile(str(self.excel_path))
        else:
            subprocess.run(["xdg-open", str(self.excel_path)])


def _load_local_config():
    """Load config.local.json from the project root (parent of this package)."""
    config_path = Path(__file__).parent.parent / "config.local.json"
    if config_path.exists():
        with open(config_path, encoding='utf-8') as f:
            return json.load(f)
    return {}


def main():
    """Main entry point"""
    print("\n" + "="*60)
    print("🏯 TEMPLE LIGHTING DONATION AUTOMATION TOOL")
    print("="*60)

    local = _load_local_config()
    saved_path = local.get("excel_file", "")
    saved_stem = local.get("year_stem", "丙午")

    # Get Excel file path
    if saved_path:
        print(f"\nSaved Excel file: {saved_path}")
        entered = input("Press Enter to use it, or type a new path: ").strip()
        excel_path = entered if entered else saved_path
    else:
        print("\nEnter the path to your Excel file:")
        print("  Example: /Users/username/Desktop/temple/点灯/丙午年初一十五点灯.xlsx")
        excel_path = input("File path: ").strip()

    if not excel_path:
        print("❌ No file path provided")
        return

    # Initialize workflow
    workflow = TempleWorkflow(excel_path, year_stem=saved_stem)

    # List available sheets
    print("\n📅 Available month sheets:")
    sheets = workflow.list_available_sheets()
    for i, sheet in enumerate(sheets, 1):
        print(f"  {i}. {sheet}")

    # Select sheet
    print("\nWhich sheet do you want to process? (e.g., '三月十五日' or just '3')")
    sheet_choice = input("Enter sheet name or number: ").strip()

    # Parse sheet choice
    if sheet_choice.isdigit():
        sheet_idx = int(sheet_choice) - 1
        if 0 <= sheet_idx < len(sheets):
            sheet_name = sheets[sheet_idx]
        else:
            print("❌ Invalid selection")
            return
    else:
        sheet_name = sheet_choice
        if sheet_name not in sheets:
            print(f"❌ Sheet '{sheet_name}' not found")
            return

    print(f"\nSelected: {sheet_name}")

    # Input method for names
    print("\nHow do you want to input names?")
    print("  1. Paste names (press Enter twice when done)")
    print("  2. Load from file")
    choice = input("Choice (1 or 2): ").strip()

    if choice == "2":
        file_path = input("Enter names file path: ").strip()
        names = workflow.load_names_from_file(file_path)
    elif choice in ("1", ""):
        names = workflow.load_names_from_input()
    else:
        # User pasted names directly at this prompt — first line was consumed as the "choice"
        names = workflow.load_names_from_input(first_line=choice)

    if not names:
        print("❌ No names provided")
        return

    print(f"\n✅ Loaded {len(names)} names")

    # Extract lunar date from sheet name
    # Examples: "三月十五日" -> "三月十五"
    lunar_date = sheet_name.replace('日', '')

    # Run the workflow
    workflow.run_workflow(sheet_name, names, lunar_date)


if __name__ == "__main__":
    main()
